import streamlit as st
import os
import textwrap
from dotenv import load_dotenv
import google.generativeai as genai
from PIL import Image
import pytesseract
from PyPDF2 import PdfReader
import pandas as pd
import openpyxl
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import logging
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_google_genai import GoogleGenerativeAIEmbeddings, ChatGoogleGenerativeAI
from langchain_community.vectorstores import FAISS
from langchain.chains.question_answering import load_qa_chain
from langchain.prompts import PromptTemplate

# Load environment variables from a .env file
load_dotenv()
os.getenv("GOOGLE_API_KEY")

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize Google Generative AI
genai.configure(api_key="AIzaSyDmzp85nBBFu2ozSIcv61cekf8IpwANWNk")

# Function to get text from image using Tesseract
def get_image_text(image):
    if image:
        return pytesseract.image_to_string(image)
    return ""

# Function to get text from Excel files
def get_excel_text(excel_files):
    text = ""
    for excel_file in excel_files:
        wb = openpyxl.load_workbook(excel_file)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        text += cell.value + " "
    logger.info("Extracted text from Excel files.")
    return text

# Function to get text from Google Sheets
def get_google_sheets_text(sheet_urls):
    text = ""
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name('sheets-430211-772619f66a93.json', scope)
    client = gspread.authorize(creds)
    
    for sheet_url in sheet_urls:
        sheet = client.open_by_url(sheet_url)
        for worksheet in sheet.worksheets():
            rows = worksheet.get_all_values()
            for row in rows:
                for cell in row:
                    if isinstance(cell, str):
                        text += cell + " "
    logger.info("Extracted text from Google Sheets.")
    return text

# Function to get text from PDF files
def get_pdf_text(pdf_docs):
    text = ""
    for pdf in pdf_docs:
        pdf_reader = PdfReader(pdf)
        for page in pdf_reader.pages:
            text += page.extract_text()
    logger.info("Extracted text from PDF files.")
    return text

# Function to split text into chunks
def get_text_chunks(text):
    text_splitter = RecursiveCharacterTextSplitter(chunk_size=10000, chunk_overlap=1000)
    chunks = text_splitter.split_text(text)
    logger.info("Split text into chunks.")
    return chunks

# Function to create and save vector store
def get_vector_store(text_chunks):
    embeddings = GoogleGenerativeAIEmbeddings(model="models/embedding-001")
    vector_store = FAISS.from_texts(text_chunks, embedding=embeddings)
    vector_store.save_local("faiss_index")
    logger.info("Created and saved vector store.")

# Function to get conversational chain
def get_conversational_chain():
    prompt_template = """
    Answer like OpenAI's ChatGPT or Google's Gemini would. Be detailed and explanative. If you are asked you provide the answer in a certain format, please do
    
    Context:\n {context}?\n
    Question: \n{question}\n

    Answer:
    """
    model = ChatGoogleGenerativeAI(model="gemini-1.5-pro", temperature=0.3)  # Updated model
    prompt = PromptTemplate(template=prompt_template, input_variables=["context", "question"])
    chain = load_qa_chain(model, chain_type="stuff", prompt=prompt)
    logger.info("Loaded QA chain.")
    return chain

# Function to handle user input
def user_input(user_question):
    embeddings = GoogleGenerativeAIEmbeddings(model="models/embedding-001")
    new_db = FAISS.load_local("faiss_index", embeddings, allow_dangerous_deserialization=True)
    docs = new_db.similarity_search(user_question)
    
    if not docs:
        return "No relevant documents found."
    
    chain = get_conversational_chain()
    response = chain.invoke({"input_documents": docs, "question": user_question})
    logger.info("Generated response to user question.")
    return response["output_text"]

# Function to get Gemini AI response
def get_gemini_response(question):
    model = genai.GenerativeModel('gemini-1.5-pro')
    response = model.generate_content(question)
    return response.text

# Initialize Streamlit app
st.set_page_config(page_title="Clair", page_icon='Colou alter logo.ico')

st.markdown(
    """
    <style>
    .sidebar-content {
        display: flex;
        flex-direction: column;
        align-items: relative;
        justify-content: flex-start;
        height: 100vh;
        padding-bottom: 40px;
    }

    .sidebar-text {
        font-size: 24px;
        font-weight: bold;
        white-space: nowrap;
        overflow: hidden;
        animation: typing 3s steps(10, end);
        margin-bottom: 10px;
    }

    .sidebar-text_tag {
        font-size: 15px;
        font-weight: normal;
        white-space: nowrap;
        overflow: hidden;
        animation: typing 3s steps(10, end);
        margin-bottom: 5px;
    }

    @keyframes typing {
        from { width: 0 }
        to { width: 100% }
    }
    </style>
    """,
    unsafe_allow_html=True
)

st.sidebar.image('Colou alter logo.jpg', width=100)
st.sidebar.markdown('<div class="sidebar-text">Clairvoyance</div>', unsafe_allow_html=True)
st.sidebar.markdown('<div class ="sidebar-text_tag">Building Sustainable Automation</div>', unsafe_allow_html=True)

# Sidebar content
with st.sidebar:
    st.title("Get Started:")
    doc_type = st.selectbox("Select Document Type", ["Drop It", "Excel/Google Sheets", "PDF", "Image"])

    if doc_type == "Excel/Google Sheets":
        st.subheader("Upload Files:")
        excel_files = st.file_uploader("Upload Excel Files", accept_multiple_files=True)
        google_sheets_urls = st.text_area("Enter Google Sheets URLs (one per line)")
        if st.button("Submit & Process"):
            with st.spinner("Processing..."):
                raw_text = ""
                if excel_files:
                    raw_text += get_excel_text(excel_files)
                if google_sheets_urls:
                    sheet_urls = google_sheets_urls.split("\n")
                    raw_text += get_google_sheets_text(sheet_urls)
                
                text_chunks = get_text_chunks(raw_text)
                get_vector_store(text_chunks)
                st.success("Done")

    elif doc_type == "PDF":
        pdf_docs = st.file_uploader("Upload your PDF Files and Click on the Submit & Process Button", accept_multiple_files=True)
        if st.button("Submit & Process"):
            with st.spinner("Processing..."):
                raw_text = get_pdf_text(pdf_docs)
                text_chunks = get_text_chunks(raw_text)
                get_vector_store(text_chunks)
                st.success("Done")

    elif doc_type == "Image":
        uploaded_file = st.file_uploader("Choose an image...", type=["jpg", "jpeg", "png"])
        if uploaded_file is not None:
            image = Image.open(uploaded_file)
            st.image(image, caption="Uploaded Image.", use_column_width=True)
            if st.button("Analyze Image"):
                with st.spinner("Processing..."):
                    image_text = get_image_text(image)
                    st.subheader("Extracted Text from Image:")
                    st.write(image_text)
                    st.success("Done")

# Main screen content
st.header("Say Hello to Clair!!🌌")
input_text = st.text_area("Input:", key="input")

if st.button("Generate"):
    if input_text:
        response = user_input(input_text)
        st.subheader("The Response is")
        st.write(response)
    else:
        st.warning("Please enter a question or input text.")

# Footer
footer = """
    <div style="text-align: center;">
        <hr>
        <p>All Copyrights Reserved Atharva Shitut @2024</p>
    </div>
"""
st.markdown(footer, unsafe_allow_html=True)
st.image('Black landscape.jpg')
